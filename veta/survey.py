import openpyxl

from veta.item import Item
from veta.respondent import Respondent
from veta.wordlist import Wordlist
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics import confusion_matrix
import os 
import json
from scipy.stats import norm

class NumpyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (np.integer, np.int64, np.int32)):
            return int(obj)
        elif isinstance(obj, (np.floating, np.float64, np.float32)):
            return float(obj)
        elif isinstance(obj, (np.ndarray,)):
            return obj.tolist()
        elif isinstance(obj, np.bool_):
            return bool(obj)
        else:
            return super(NumpyEncoder, self).default(obj)

class Survey:
    """
    A class representing a single 
    ...

    Attributes
    ----------
    items: list
        a list containing all of the LEAS items belonging to the Respondent.
    
    Methods
    -------
    __str__()
        handles conversion of the respondent to a string object for display
    
    """
    def __init__(self, wordlist_file=None) -> None:
        self.respondents = []
        self.wordlist = None
        if isinstance(wordlist_file,str):
            wordlist = Wordlist(wordlist_file)
            self.add_wordlist(wordlist)

        self.cols = [0,1,2]
        self.num_item_cols = 0
        self.summary = {}
        self.header = np.array(["ID", "Self", "Other"])
        return

    def __str__(self) -> str:
        ret = ""
        for respondent in self.respondents:
            # ret += "Respondent ID {}:\n\n".format(respondent.id)
            ret += str(respondent) + '\n\n'
        return ret.rstrip()

    def from_vertical_layout(self, data):

        self.header = data[0,:]
        data = data[1:,]
        self.data = data
        id_col, self_col, other_col = self.cols[:3]

        res = Respondent()
        self.add_respondent(res)
        
        for i in range(data.shape[0]):
            userid = data[i,id_col]
            if isinstance(userid,float) and np.isnan(userid):
                for col in self.cols[3:3+self.num_item_cols ]:
                    total = 0
                    for item in res.items:
                        total += item.scores[self.header[col]]
                    #print(self.header[col], total)
                    res.add_additional_info(self.header[col], total)
                    
                for col in self.cols[3+self.num_item_cols:]:
                    #print(self.header[col], data[i,col])
                    res.add_additional_info(self.header[col], data[i,col])
                res = Respondent()
                self.add_respondent(res)
                
            else:
                res.userid=str(userid)
                self_sentence = data[i,self_col]
                other_sentence = data[i,other_col]
                if isinstance(self_sentence,float) and np.isnan(self_sentence):
                    self_sentence = ""
                if isinstance(other_sentence,float) and np.isnan(other_sentence):
                    other_sentence = ""
                
                item = res.add_item(self_sentence,other_sentence)
                for col in self.cols[3:3+self.num_item_cols ]:
                    #print(self.header[col], col,  data[i,col])
                    item.add_additional_info(self.header[col], data[i,col])
                
        if len(res.items) == 0:
            self.respondents.remove(res)
        return

    def from_horizontal_layout(self, data):

        self.header = data[0,:]
        data = data[1:,]
        self.data = data
        id_col, self_col, other_col = self.cols[:3]

        #Loop through rows of the data
        for i in range(data.shape[0]):
            #Extract the userID
            userid = data[i,id_col]
            if isinstance(userid,float) and np.isnan(userid):
                res = Respondent()
            else:
                res = Respondent(userid=str(userid))
            self.add_respondent(res)
            #loop through data columns, starting with self_col
            #Assumes LEAS data is continuous
            for j in range(self_col, data.shape[1],2):

                self_sentence = data[i,j]
                other_sentence = data[i,j+1]
                if isinstance(self_sentence,float) and np.isnan(self_sentence):
                    self_sentence = ""
                if isinstance(other_sentence,float) and np.isnan(other_sentence):
                    other_sentence = ""
                res.add_item(self_sentence,other_sentence)
            for col in self.cols[3+self.num_item_cols:]:
                res.add_additional_info(self.header[col], data[i,col])
        return        

    def from_file(self, filename, layout='vertical'):

        # Get the file extension
        file_extension = os.path.splitext(filename)[1]
        
        # Check the file extension and read the file accordingly
        if file_extension == '.csv':
            data = np.array(pd.read_csv(filename, header=None))
        elif file_extension in ['.xls', '.xlsx']:
            data = np.array(pd.read_excel(filename, header=None, engine='openpyxl'))
        elif file_extension in ['.json']:
            self.from_json(filename)
            return
        else:
            raise ValueError(f"Unsupported file extension: {file_extension}")

        if str(layout).lower() == "vertical":
            self.from_vertical_layout(data)
        elif str(layout).lower() == "horizontal":
            self.from_horizontal_layout(data)

        self.add_wordlist(self.wordlist)

        return
    
    def configure_columns(self, id_col, self_col, other_col, per_item_cols=[], per_res_cols=[]):
        self.cols = [id_col,self_col,other_col]
        for col in per_item_cols:
            self.cols.append(col)
        self.num_item_cols = len(per_item_cols)
        for col in per_res_cols:
            self.cols.append(col)
        return

    def add_respondent(self, respondent):
        self.respondents.append(respondent)
        if not (self.wordlist is None):
            respondent.add_wordlist(self.wordlist)
        return

    def score(self,*modules):

        for respondent in self.respondents:
            respondent.score(*modules)

    def compute_summary(self, percentiles=False):

        for respondent in self.respondents:
            for key in respondent.totals.keys():
                self.summary[key] = []
        
        for respondent in self.respondents:
            for key in respondent.totals.keys():
                self.summary[key].append(respondent.totals[key])
        
        for key in self.summary.keys():
            self.summary[key] = np.array(self.summary[key], dtype=float)

        if percentiles:
            if "3345plus" in self.summary.keys():
                # Convert each value to a percentile
                self.summary['20-item-percentile']  = norm.cdf(self.summary["3345plus"], 
                                       loc=70, 
                                       scale=7) * 100
                                # Convert each value to a percentile
                self.summary['10-item-percentile']  = norm.cdf(self.summary["3345plus"], 
                                       loc=35, 
                                       scale=5) * 100


    def add_wordlist(self, wordlist: Wordlist):
        self.wordlist = wordlist
        for respondent in self.respondents:
            respondent.add_wordlist(wordlist)

    def save(self, filename):
        # Get the file extension
        file_extension = os.path.splitext(filename)[1]

        if file_extension in ['.json']:
            with open(filename, 'w') as f:
                json.dump(self.to_json(), f, indent=2, cls=NumpyEncoder)
            return

        #Open a new excel workbook
        wb = openpyxl.Workbook()
        ws1 = wb.active
        # data = self.data
        num_cols = 3

        #Set the header row
        for j in range(num_cols):
            ws1.cell(1, column=j+1).value = self.header[self.cols[j]]

        #Get the module names that have been run and sort them
        respondent_modules = list(self.respondents[0].totals.keys())
        respondent_modules.sort()

        #Write the module names to the columns
        for j in range(len(respondent_modules)):
            ws1.cell(1, column=num_cols+j+1).value = respondent_modules[j]

        #Copy the first three columns (Person Self Other)
        # for i in range(data.shape[0]):
        #     for j in range(num_cols):
        #         ws1.cell(row=i+2, column=j+1).value = data[i,j]
        r = 2
        for res in self.respondents:
            for item in res.items:
                if isinstance(res.userid,str):
                    ws1.cell(row=r, column=1).value = res.userid
                else:
                    ws1.cell(row=r, column=1).value = res.userid
                ws1.cell(row=r, column=2).value = item.self_sentence
                ws1.cell(row=r, column=3).value = item.other_sentence
                r += 1
            r +=1

        #Output all of the results
        current_row = 1
        for respondent in self.respondents:
            full_data = respondent.to_array()
            for i in range(full_data.shape[0]):
                for j in range(full_data.shape[1]):
                    ws1.cell(row=current_row+1, column=j+num_cols+1).value = full_data[i,j]
                current_row += 1

        if file_extension == '.csv':
            # Convert the first sheet of the workbook to a pandas DataFrame
            ws = wb.active  # or specify the sheet by wb[sheetname]
            data = ws.values
            # Create a DataFrame from the sheet data
            df = pd.DataFrame(data)
            # Save the DataFrame as a CSV file
            csv_filename = filename if filename.endswith('.csv') else filename + '.csv'
            df.to_csv(csv_filename, index=False, header=False)
        elif file_extension in ['.xls', '.xlsx']:
            # Save as the original workbook format (Excel)
            wb.save(filename=filename)
        else:
            raise ValueError(f"Unsupported file extension: {file_extension}")


    
    def plot_matrix(self, ids):

        if isinstance(ids, int) and ids == -1:
            ids = list(self.summary.keys())

        if isinstance(ids, list):
            assert len(ids) >= 2
            for i in range(len(ids)):
                assert isinstance(ids[i], str)

        height, width = len(ids)-1, len(ids)-1

        fig, ax = plt.subplots(nrows=height, ncols=width, figsize = (5*height,5*width))
        if height == 1 and width == 1:
            ax = [[ax]]
        
        for i in range(height):
            row = ax[i]
            for j in range(width):
                col = row[j]
                if i <= j:
                    x = self.summary[ids[i]]
                    
                    y = self.summary[ids[j+1]]
                    y = y[~np.isnan(x)]
                    x = x[~np.isnan(x)]
                    x = x[~np.isnan(y)]
                    y = y[~np.isnan(y)]
                    
                    corr = np.corrcoef(x, y)[1,0]
                    z = np.polyfit(x, y, 1)

                    col.plot(x, y, '.', color = 'k', alpha = 0.5, label = r"$\rho = ${:.2f}".format(corr))
                    col.plot(x, z[0]*x + z[1], '-', color = 'r', alpha = 0.3)
                    col.set_xlabel(ids[i],size = 16)
                    col.set_ylabel(ids[j+1],size = 16)
                    col.grid()
                    col.legend(loc = 0)
                else:
                    col.set_axis_off()

        fig.tight_layout()
        plt.show()

    def plot_confusion(self, keys):

        if len(keys) != 2 or keys[0] not in self.summary.keys() or keys[1] not in self.summary.keys():
            raise ValueError("Must give two proper, per-item survey keys")

        confusion = dict()
        for _id in keys:
            confusion[_id] = []
        for r in self.respondents:
            for item in r.items:
                for _id in keys:
                    confusion[_id].append(item.scores[_id])

        # Calculate the confusion matrix
        cm_svm_tf = confusion_matrix(confusion[keys[0]], confusion[keys[1]])

        # Create a dataframe from the confusion matrix for better visualization
        confusion_matrix_df_svm_tf = pd.DataFrame(cm_svm_tf)

        plt.figure(figsize=(10, 8))
        sns.heatmap(confusion_matrix_df_svm_tf, annot=True, cmap='Blues', fmt='d')
        plt.xlabel(keys[1])
        plt.ylabel(keys[0])
        plt.title(f'Confusion Matrix Between {keys[0]} & {keys[1]}')
        plt.show() 

    def to_json(self):
        data = []
        for respondent in self.respondents:
            respondent_data = {
                'userid': respondent.userid,
                'totals': respondent.totals,
                'items': []
            }
            for item in respondent.items:
                item_data = {
                    'self_sentence': item.self_sentence,
                    'other_sentence': item.other_sentence,
                    'scores': item.scores
                }
                respondent_data['items'].append(item_data)
            data.append(respondent_data)
        return data

    def from_json(self, filename):

        try:
            with open(filename, 'r') as f:
                data = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError) as e:
            print(f"Error loading JSON data: {e}")
            return

        self.respondents = []

        for respondent_data in data:
            respondent = Respondent()
            respondent.userid = respondent_data.get('userid', '')
            respondent.totals = respondent_data.get('totals', {})
            respondent.items = []

            for item_data in respondent_data.get('items', []):

                item = Item(item_data.get('self_sentence', ''),
                           item_data.get('other_sentence', ''))
                item.scores = item_data.get('scores', {})

                respondent.items.append(item)

            self.respondents.append(respondent)

        self.add_wordlist(self.wordlist)
        return
