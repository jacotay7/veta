import openpyxl

# Load data from XLSX file
data_col1 = []
data_col2 = []
workbook = openpyxl.load_workbook('en(1).xlsx')
sheet = workbook.active

# Move every third row starting from the 1st row to the first column
# Move every third row starting from the 2nd row to the second column
for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
    if i % 3 == 1:  # Check every third row starting from the 1st row
        data_col1.append(row[0])  # Append data to the first column
    elif i % 3 == 2:  # Check every third row starting from the 2nd row
        data_col2.append(row[0])  # Append data to the second column

# Create Excel workbook and sheet
workbook_output = openpyxl.Workbook()
sheet_output = workbook_output.active

# Write data to the first column
for i, value in enumerate(data_col1, start=1):  # Start counter at 1
    sheet_output.cell(row=i, column=1).value = value

# Write data to the second column
for i, value in enumerate(data_col2, start=1):  # Start counter at 1
    sheet_output.cell(row=i, column=2).value = value

# Save the Excel workbook
workbook_output.save('output.xlsx')
