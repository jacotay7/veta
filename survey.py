import openpyxl
from item import *
from respondent import *
from wordlist import *
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

class Survey:

    def __init__(self, wordlist_file=None) -> None:
        self.respondents = []
        self.wordlist = None
        if isinstance(wordlist_file,str):
            wordlist = Wordlist(wordlist_file)
            self.add_wordlist(wordlist)

        self.cols = [0,1,2]
        self.num_item_cols = 0
        self.summary = {}
        return

    def __str__(self) -> str:
        ret = ""
        for respondent in self.respondents:
            ret += "Respondent ID {}:\n\n".format(respondent.id)
            ret += str(respondent) + '\n\n'
        return ret.rstrip()

    def from_file(self, filename):

        data = np.array(pd.read_excel(filename, header = None))
        self.header = data[0,:]#np.array([data[0,col] for col in self.cols])

        data = data[1:,]
        self.data = data
        id_col, self_col, other_col = self.cols[:3]

        res = Respondent()
        self.add_respondent(res)
        for i in range(data.shape[0]):
            if isinstance(data[i,id_col],float) and np.isnan(data[i,id_col]):
                for col in self.cols[3:3+self.num_item_cols ]:
                    total = 0
                    for item in res.items:
                        total += item.scores[self.header[col]]
                    #print(self.header[col], total)
                    res.add_additional_info(self.header[col], total)
                    
                for col in self.cols[3+self.num_item_cols:]:
                    #print(self.header[col], data[i,col])
                    res.add_additional_info(self.header[col], data[i,col])
                res = Respondent()
                self.add_respondent(res)
                
            else:
                self_sentence = data[i,self_col]
                other_sentence = data[i,other_col]
                if isinstance(self_sentence,float) and np.isnan(self_sentence):
                    self_sentence = ""
                if isinstance(other_sentence,float) and np.isnan(other_sentence):
                    other_sentence = ""
                item = res.add_item(self_sentence,other_sentence)
                for col in self.cols[3:3+self.num_item_cols ]:
                    #print(self.header[col], col,  data[i,col])
                    item.add_additional_info(self.header[col], data[i,col])
        self.respondents.remove(res)
        """for col in self.cols[3:3+self.num_item_cols ]:
            total = 0
            for item in res.items:
                total += item.scores[self.header[col]]
            print(self.header[col], total)
            res.add_additional_info(self.header[col], total)
        for col in self.cols[3+self.num_item_cols:]:
            res.add_additional_info(self.header[col], data[i,col])"""
        
        self.add_wordlist(self.wordlist)
        return
    
    def configure_columns(self, id_col, self_col, other_col, per_item_cols, per_res_cols):
        self.cols = [id_col,self_col,other_col]
        for col in per_item_cols:
            self.cols.append(col)
        self.num_item_cols = len(per_item_cols)
        for col in per_res_cols:
            self.cols.append(col)
        return

    def add_respondent(self, respondent):
        self.respondents.append(respondent)
        return

    def score(self,*modules):

        for respondent in self.respondents:
            respondent.score(*modules)

    def compute_summary(self):

        for respondent in self.respondents:
            for key in respondent.totals.keys():
                self.summary[key] = []
        
        for respondent in self.respondents:
            for key in respondent.totals.keys():
                self.summary[key].append(respondent.totals[key])
        
        for key in self.summary.keys():
            self.summary[key] = np.array(self.summary[key], dtype=float)

    def add_wordlist(self, wordlist):
        self.wordlist = wordlist
        for respondent in self.respondents:
            respondent.add_wordlist(wordlist)

    def save(self, filename):

        wb = openpyxl.Workbook()
        #print(self.data[0,:])
        ws1 = wb.active
        data = self.data
        num_cols = 3

        for j in range(num_cols):
            ws1.cell(1, column=j+1).value = self.header[j]

        respondent_modules = list(self.respondents[0].totals.keys())
        respondent_modules.sort()

        for j in range(len(respondent_modules)):
            ws1.cell(1, column=num_cols+j+1).value = respondent_modules[j]

        for i in range(data.shape[0]):
            for j in range(num_cols):
                ws1.cell(row=i+2, column=j+1).value = data[i,j]

        current_row = 1
        for respondent in self.respondents:
            full_data = respondent.to_array()
            for i in range(full_data.shape[0]):
                for j in range(full_data.shape[1]):
                    ws1.cell(row=current_row+1, column=j+num_cols+1).value = full_data[i,j]
                current_row += 1

        wb.save(filename = filename)

    
    def plot_matrix(self, ids):

        if ids == -1:
            ids = list(self.summary.keys())

        fig, ax = plt.subplots(nrows=len(ids), ncols=len(ids), figsize = (20,20))
        fig.tight_layout()
        for i in range(len(ids)):
            row = ax[i,:]
            for j in range(len(row)):
                col = row[j]
                if i <= j:
                    x = self.summary[ids[i]]
                    
                    y = self.summary[ids[j]]
                    y = y[~np.isnan(x)]
                    x = x[~np.isnan(x)]
                    x = x[~np.isnan(y)]
                    y = y[~np.isnan(y)]
                    
                    corr = np.corrcoef(x, y)[1,0]
                    z = np.polyfit(x, y, 1)

                    col.plot(x, y, '.', color = 'k', alpha = 0.5, label = r"$\rho = ${:.2f}".format(corr))
                    col.plot(x, z[0]*x + z[1], '-', color = 'r', alpha = 0.3)
                    col.set_xlabel(ids[i],size = 10)
                    col.set_ylabel(ids[j],size = 10)
                    col.grid()
                    col.legend(loc = 0)
                else:
                    col.set_axis_off()


        plt.show()

